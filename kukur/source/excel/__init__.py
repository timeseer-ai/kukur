"""Read Excel files to Apache Arrow tables."""

# SPDX-FileCopyrightText: 2025 Timeseer.AI
# SPDX-License-Identifier: Apache-2.0

import pyarrow as pa

from kukur.inspect import DataOptions

try:
    import openpyxl

    HAS_OPENPYXL = True

except ImportError:
    HAS_OPENPYXL = False


from kukur.exceptions import MissingModuleException


def parse_excel(readable, sheet_name: str, options: DataOptions | None) -> pa.Table:
    """Parse an Excel file to a PyArrow Table."""
    if not HAS_OPENPYXL:
        raise MissingModuleException("openpyxl", "excel")

    workbook = openpyxl.load_workbook(readable, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]

    data = list(worksheet.values)
    workbook.close()

    if len(data) == 0:
        return _to_pyarrow([], [])
    if options is not None and options.excel_header_row:
        headers = data[0]
        rows = data[1:]
    else:
        headers = [str(i) for i in range(len(data[0]))] if len(data) > 0 else []
        rows = data

    return _to_pyarrow([str(header) for header in headers], rows)


def list_sheets(readable) -> list[str]:
    """List the sheets in the Excel file."""
    if not HAS_OPENPYXL:
        raise MissingModuleException("openpyxl", "excel")
    workbook = openpyxl.load_workbook(readable, read_only=True)
    sheetnames = workbook.sheetnames
    workbook.close()
    return sheetnames


def _to_pyarrow(headers: list[str], rows: list[list]) -> pa.Table:
    """Convert raw excel to a PyArrow table."""
    arrays = []
    columns = list(zip(*rows, strict=True))
    for column in columns:
        try:
            arrays.append(pa.array(column))
        except pa.lib.ArrowException:
            arrays.append(
                pa.array(
                    [str(value) if value is not None else None for value in column],
                    type=pa.string(),
                )
            )

    return pa.table(arrays, names=headers)
